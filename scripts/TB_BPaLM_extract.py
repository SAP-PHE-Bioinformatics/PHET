import json
import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Genes of interest
target_genes = {
    "atpE", "Rv1979c", "pepQ", "mmpL5", "mmpR5", "ddn",
    "fgd1", "fbiB", "fbiC", "fbiD", "rplC", "rrl"
}

# Sections that contain mutations
sections = {
    "dr_variants": "Resistance variants report",
    "other_variants": "Other variants report",
    "soft_fail_variants": "QC failed variants report"
}

# Map confidence to action
def determine_action(section_name, confidence, change):
    if confidence is None or str(confidence).strip() == "":
        return "If a frameshift or deletion, this may be listed different in the WHO catalogue, otherwise if mut. found at a high frequency, investigate from literature"
    if section_name == "QC failed variants report":
        return "Failed QC"
    elif confidence in {"Assoc w R", "Assoc w R - Interim"}:
        return "Report"
    elif confidence == "Uncertain significance":
        return "Further investigation required"
    elif confidence in {"Not assoc w R", "Not assoc w R - Interim"}:
        return "Not reportable"
    return ""

# Map actions to colors
action_colors = {
    "Report": "90EE90",  # Light green
    "Further investigation required": "FFFF00",  # Yellow
    "Not reportable": "D3D3D3",  # Light gray
    "Failed QC": "FF7074"  # Light red
}

def main():
    parser = argparse.ArgumentParser(
        description="Extract drug resistance of new line drug mutations by gene from JSON report"
    )
    parser.add_argument("input_file", help="Path to input JSON report")
    args = parser.parse_args()

    input_path = Path(args.input_file)

    # Check file extension
    if input_path.suffix.lower() != ".json":
        print(f"❌ Error: The input file must be a JSON file, not '{input_path.suffix}'. Please provide a .json file.")
        return

    base_name = input_path.stem.split(".")[0]
    output_file = input_path.with_name(f"{base_name}_NewLineDrugMutations.xlsx")

    with open(args.input_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    found_genes = set()

    wb = Workbook()
    ws = wb.active
    ws.title = "Mutations Report"

    # Header
    header = [
        "Section", "Chrom", "Pos", "Ref", "Alt", "Depth", "Freq",
        "Gene ID", "Gene Name", "Variant Type", "Change",
        "Nucleotide Change", "Protein Change", "Drug", "Confidence", "Comment", "Action",
        "WHO catalogue format", "Reporting format"
    ]
    ws.append(header)

    for key, section_name in sections.items():
        if key not in data:
            continue
        for var in data[key]:
            gene = var.get("gene_name")
            if gene in target_genes:
                found_genes.add(gene)

                annotations = var.get("annotation", [])
                change_val = var.get("change") or ""

                if annotations:
                    for ann in annotations:
                        confidence = ann.get("confidence")
                        freq_val = var.get("freq")

                        # Format frequency to two decimal places
                        try:
                            freq_val_formatted = f"{float(freq_val):.2f}" if freq_val is not None else ""
                        except (ValueError, TypeError):
                            freq_val_formatted = ""

                        action = determine_action(section_name, confidence, change_val)
                        who_name = f"{gene}_{change_val}" if action in {"Report", "Further investigation required"} else ""
                        reporting_format = f"{gene} {change_val} ({freq_val_formatted})" if action in {"Report", "Further investigation required"} else ""

                        row = [
                            section_name,
                            var.get("chrom", ""),
                            var.get("pos", ""),
                            var.get("ref", ""),
                            var.get("alt", ""),
                            var.get("depth", ""),
                            var.get("freq", ""),
                            var.get("gene_id", ""),
                            gene,
                            var.get("type", ""),
                            change_val,
                            var.get("nucleotide_change", ""),
                            var.get("protein_change", ""),
                            ann.get("drug", ""),
                            confidence or "",
                            ann.get("comment", ""),
                            action,
                            who_name,
                            reporting_format
                        ]
                        ws.append(row)

                        # Color fill for Action column
                        action_col_idx = header.index("Action") + 1
                        if action in action_colors:
                            ws.cell(row=ws.max_row, column=action_col_idx).fill = PatternFill(
                                start_color=action_colors[action],
                                end_color=action_colors[action],
                                fill_type="solid"
                            )
                else:
                    freq_val = var.get("freq")
                    try:
                        freq_val_formatted = f"{float(freq_val):.2f}" if freq_val is not None else ""
                    except (ValueError, TypeError):
                        freq_val_formatted = ""

                    action = determine_action(section_name, None, change_val)
                    who_name = f"{gene}_{change_val}" if action in {"Report", "Further investigation required"} else ""
                    reporting_format = f"{gene} {change_val} ({freq_val_formatted})" if action in {"Report", "Further investigation required"} else ""

                    ws.append([
                        section_name,
                        var.get("chrom", ""),
                        var.get("pos", ""),
                        var.get("ref", ""),
                        var.get("alt", ""),
                        var.get("depth", ""),
                        var.get("freq", ""),
                        var.get("gene_id", ""),
                        gene,
                        var.get("type", ""),
                        change_val,
                        var.get("nucleotide_change", ""),
                        var.get("protein_change", ""),
                        "", "", "", action, who_name, reporting_format
                    ])

    # Add summary section
    ws.append([])
    ws.append(["Summary of target genes:"])
    ws.append([f"✅ Found: {', '.join(sorted(found_genes)) if found_genes else 'None'}"])
    missing_genes = target_genes - found_genes
    ws.append([f"❌ Not Detected: {', '.join(sorted(missing_genes)) if missing_genes else 'None'}"])

    wb.save(output_file)
    print(f"Filtered mutations written to {output_file}")


if __name__ == "__main__":
    main()